"""
VCD Waveform Converter - Converts VCD to CSV, JSON, or Excel.
Usage: python vcd_converter.py [input.vcd] [options]
Options: --csv/--json/--excel, --hex/--int/--signed/--smag/--bin, --us/--ns/--ps
"""

import sys
import re
import json
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent.resolve()
VCD_INPUT_DIR = SCRIPT_DIR / "vcd_output"
OUTPUT_DIR = SCRIPT_DIR / "converted_output"

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    HAS_TK = True
except ImportError:
    HAS_TK = False


# ============================================================
# VCD Parser
# ============================================================

def parse_vcd(filename):
    """Parse VCD file. Returns (signals, changes, metadata)."""
    with open(filename, 'r') as f:
        content = f.read()
    
    # Metadata
    date_match = re.search(r'\$date\s+(.+?)\s+\$end', content, re.DOTALL)
    version_match = re.search(r'\$version\s+(.+?)\s+\$end', content, re.DOTALL)
    
    metadata = {
        'date': date_match.group(1).strip() if date_match else '',
        'version': version_match.group(1).strip() if version_match else '',
        'filename': Path(filename).name
    }
    
    # Signals: $var wire 4 ! A [3:0] $end
    signals = {}
    pattern = r'\$var\s+(\w+)\s+(\d+)\s+(\S+)\s+(\w+)(?:\s+\[[\d:]+\])?\s+\$end'
    for match in re.finditer(pattern, content):
        var_type, width, var_id, name = match.groups()
        signals[var_id] = {
            'name': name,
            'width': int(width),
            'type': var_type
        }
    
    # Value changes
    changes = []
    current_time = 0
    
    for line in content.split('\n'):
        line = line.strip()
        if not line:
            continue
        
        if line.startswith('#'):
            try:
                current_time = int(line[1:])
            except ValueError:
                pass
        elif line.startswith('b'):
            match = re.match(r'b([01xXzZ]+)\s+(\S+)', line)
            if match:
                value, var_id = match.groups()
                if var_id in signals:
                    changes.append((current_time, var_id, value))
        elif len(line) >= 2 and line[0] in '01xXzZ':
            var_id = line[1:]
            if var_id in signals:
                changes.append((current_time, var_id, line[0]))
    
    return signals, changes, metadata


def build_timeline(signals, changes):
    """Build complete timeline with all signal values at each timestamp."""
    
    current_values = {vid: '0' for vid in signals}
    timestamps = sorted(set(t for t, _, _ in changes))
    
    # Group changes by time
    by_time = {}
    for t, vid, val in changes:
        if t not in by_time:
            by_time[t] = {}
        by_time[t][vid] = val
    
    # Build rows
    rows = []
    for t in timestamps:
        if t in by_time:
            current_values.update(by_time[t])
        rows.append((t, dict(current_values)))
    
    return rows


# ============================================================
# Value Formatting
# ============================================================

def format_value(binary_str, fmt='hex', width=None):
    """Format binary string to hex/int/signed/smag/bin."""
    if fmt == 'bin':
        return binary_str
    if 'x' in binary_str.lower() or 'z' in binary_str.lower():
        return binary_str
    
    try:
        bits = len(binary_str)
        unsigned_val = int(binary_str, 2)
        
        if fmt == 'hex':
            return format(unsigned_val, 'X')
        elif fmt == 'int':
            return str(unsigned_val)
        elif fmt == 'signed':
            if binary_str[0] == '1':
                return str(unsigned_val - (1 << bits))
            return str(unsigned_val)
        elif fmt == 'smag':
            if bits == 1:
                return str(unsigned_val)
            magnitude = int(binary_str[1:], 2) if len(binary_str) > 1 else 0
            return str(-magnitude) if binary_str[0] == '1' else str(magnitude)
        return str(unsigned_val)
    except ValueError:
        return binary_str


def time_divisor(unit):
    """Get divisor for time unit (VCD is in picoseconds)."""
    return {'ps': 1, 'ns': 1000, 'us': 1000000, 'ms': 1000000000}.get(unit, 1000000)


# ============================================================
# Export Functions
# ============================================================

def ensure_output_dir():
    """Create output directory if it doesn't exist."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    return OUTPUT_DIR


def get_output_path(filename):
    """Get output path in the converted_output folder."""
    ensure_output_dir()
    return OUTPUT_DIR / Path(filename).name


def export_csv(signals, rows, output_file, time_unit='us', value_fmt='hex'):
    """Export to CSV format."""
    
    divisor = time_divisor(time_unit)
    names = [signals[vid]['name'] for vid in sorted(signals.keys())]
    
    with open(output_file, 'w') as f:
        f.write(f"Time({time_unit})," + ",".join(names) + "\n")
        
        for t, values in rows:
            time_str = str(t / divisor)
            vals = [format_value(values[vid], value_fmt) for vid in sorted(signals.keys())]
            f.write(time_str + "," + ",".join(vals) + "\n")
    
    return len(rows)


def export_json(signals, rows, output_file, metadata, time_unit='us', value_fmt='hex'):
    """Export to JSON format."""
    
    divisor = time_divisor(time_unit)
    signal_list = [{'id': vid, **signals[vid]} for vid in sorted(signals.keys())]
    
    data = {
        'metadata': metadata,
        'time_unit': time_unit,
        'signals': signal_list,
        'data': []
    }
    
    for t, values in rows:
        row = {'time': t / divisor}
        for vid in sorted(signals.keys()):
            row[signals[vid]['name']] = format_value(values[vid], value_fmt)
        data['data'].append(row)
    
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2)
    
    return len(rows)


def get_numeric_value(binary_str, fmt):
    """Convert binary to numeric value for Excel. Returns int or None for hex/bin/unknown."""
    if fmt not in ('int', 'signed', 'smag') or 'x' in binary_str.lower() or 'z' in binary_str.lower():
        return None
    
    try:
        bits = len(binary_str)
        val = int(binary_str, 2)
        
        if fmt == 'int':
            return val
        elif fmt == 'signed':
            return val - (1 << bits) if binary_str[0] == '1' else val
        elif fmt == 'smag':
            if bits == 1:
                return val
            mag = int(binary_str[1:], 2) if len(binary_str) > 1 else 0
            return -mag if binary_str[0] == '1' else mag
    except ValueError:
        return None


def export_excel(signals, rows, output_file, time_unit='us', value_fmt='hex'):
    """Export to Excel. Numeric formats (int/signed/smag) stored as numbers for graphing."""
    
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("openpyxl not found. Installing automatically...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            import openpyxl  # type: ignore
            print("openpyxl installed successfully!")
        except Exception as e:
            raise ImportError(f"Failed to install openpyxl: {e}\nPlease run: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waveform"
    
    divisor = time_divisor(time_unit)
    names = [signals[vid]['name'] for vid in sorted(signals.keys())]
    
    # Header style
    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Header
    headers = [f"Time({time_unit})"] + names
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data
    for row_num, (t, values) in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=t / divisor)
        for col, vid in enumerate(sorted(signals.keys()), 2):
            binary_str = values[vid]
            num_val = get_numeric_value(binary_str, value_fmt)
            ws.cell(row=row_num, column=col, value=num_val if num_val is not None else format_value(binary_str, value_fmt))
    
    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 20)
    
    wb.save(output_file)
    return len(rows)


# ============================================================
# GUI
# ============================================================

class ConverterGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("VCD Waveform Converter")
        
        # Window size
        width, height = 800, 400
        
        # Center on screen
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        x = (screen_w - width) // 2
        y = (screen_h - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.resizable(True, True)
        
        self.vcd_file = tk.StringVar()
        self.output_file = tk.StringVar()
        self.format_var = tk.StringVar(value='csv')
        self.value_fmt = tk.StringVar(value='hex')
        self.time_unit = tk.StringVar(value='us')
        
        self.build_ui()
        
        # Auto-load waveform.vcd if it exists
        default_vcd = VCD_INPUT_DIR / "waveform.vcd"
        if default_vcd.exists():
            self.vcd_file.set(str(default_vcd))
            # Set default output path in converted_output folder
            ext = {'csv': '.csv', 'json': '.json', 'excel': '.xlsx'}[self.format_var.get()]
            out_path = get_output_path("waveform" + ext)
            self.output_file.set(str(out_path))
    
    def build_ui(self):
        # Configure style
        style = ttk.Style()
        style.configure('Convert.TButton', font=('Segoe UI', 11, 'bold'), padding=(20, 10))
        style.configure('TLabel', font=('Segoe UI', 10))
        style.configure('TRadiobutton', font=('Segoe UI', 10))
        style.configure('TEntry', font=('Segoe UI', 10))
        
        # Main frame with padding
        main = ttk.Frame(self.root, padding="20")
        main.grid(row=0, column=0, sticky="nsew")
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main.columnconfigure(1, weight=1)
        
        row = 0
        
        # Input file
        ttk.Label(main, text="VCD File:").grid(row=row, column=0, sticky="w", pady=8)
        input_frame = ttk.Frame(main)
        input_frame.grid(row=row, column=1, sticky="ew", pady=8)
        input_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(input_frame, textvariable=self.vcd_file).grid(row=0, column=0, sticky="ew")
        ttk.Button(input_frame, text="Browse...", command=self.browse_input).grid(row=0, column=1, padx=(8,0))
        
        row += 1
        
        # Output file
        ttk.Label(main, text="Output:").grid(row=row, column=0, sticky="w", pady=8)
        output_frame = ttk.Frame(main)
        output_frame.grid(row=row, column=1, sticky="ew", pady=8)
        output_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(output_frame, textvariable=self.output_file).grid(row=0, column=0, sticky="ew")
        ttk.Button(output_frame, text="Browse...", command=self.browse_output).grid(row=0, column=1, padx=(8,0))
        
        row += 1
        
        # Separator
        ttk.Separator(main, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky="ew", pady=15)
        row += 1
        
        # Output format
        ttk.Label(main, text="Format:").grid(row=row, column=0, sticky="w", pady=8)
        fmt_frame = ttk.Frame(main)
        fmt_frame.grid(row=row, column=1, sticky="w", pady=8)
        
        ttk.Radiobutton(fmt_frame, text="CSV", variable=self.format_var, value='csv', command=self.update_output_ext).pack(side="left")
        ttk.Radiobutton(fmt_frame, text="JSON", variable=self.format_var, value='json', command=self.update_output_ext).pack(side="left", padx=15)
        ttk.Radiobutton(fmt_frame, text="Excel", variable=self.format_var, value='excel', command=self.update_output_ext).pack(side="left")
        
        row += 1
        
        # Value format
        ttk.Label(main, text="Values:").grid(row=row, column=0, sticky="w", pady=8)
        val_frame = ttk.Frame(main)
        val_frame.grid(row=row, column=1, sticky="w", pady=8)
        
        ttk.Radiobutton(val_frame, text="Hex", variable=self.value_fmt, value='hex').pack(side="left")
        ttk.Radiobutton(val_frame, text="Unsigned", variable=self.value_fmt, value='int').pack(side="left", padx=10)
        ttk.Radiobutton(val_frame, text="Signed", variable=self.value_fmt, value='signed').pack(side="left", padx=10)
        ttk.Radiobutton(val_frame, text="Sign-Mag", variable=self.value_fmt, value='smag').pack(side="left", padx=10)
        ttk.Radiobutton(val_frame, text="Binary", variable=self.value_fmt, value='bin').pack(side="left")
        
        row += 1
        
        # Time unit
        ttk.Label(main, text="Time:").grid(row=row, column=0, sticky="w", pady=8)
        time_frame = ttk.Frame(main)
        time_frame.grid(row=row, column=1, sticky="w", pady=8)
        
        ttk.Radiobutton(time_frame, text="μs", variable=self.time_unit, value='us').pack(side="left")
        ttk.Radiobutton(time_frame, text="ns", variable=self.time_unit, value='ns').pack(side="left", padx=15)
        ttk.Radiobutton(time_frame, text="ps", variable=self.time_unit, value='ps').pack(side="left")
        
        row += 1
        
        # Separator
        ttk.Separator(main, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky="ew", pady=15)
        row += 1
        
        # Convert button - larger and centered
        btn_frame = ttk.Frame(main)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=15)
        ttk.Button(btn_frame, text="  Convert  ", style='Convert.TButton', command=self.convert).pack()
        
        row += 1
        
        # Status
        self.status = ttk.Label(main, text="Select a VCD file to convert", foreground="gray")
        self.status.grid(row=row, column=0, columnspan=2, pady=8)
    
    def update_output_ext(self):
        """Update output file extension when format changes."""
        output = self.output_file.get()
        if output:
            ext = {'csv': '.csv', 'json': '.json', 'excel': '.xlsx'}[self.format_var.get()]
            # Replace extension
            new_output = str(Path(output).with_suffix(ext))
            self.output_file.set(new_output)
    
    def browse_input(self):
        # Default to vcd_output folder if it exists
        initial_dir = str(VCD_INPUT_DIR) if VCD_INPUT_DIR.exists() else None
        filename = filedialog.askopenfilename(
            title="Select VCD file",
            initialdir=initial_dir,
            filetypes=[("VCD files", "*.vcd"), ("All files", "*.*")]
        )
        if filename:
            self.vcd_file.set(filename)
            # Auto-set output filename in converted_output folder
            ext = {'csv': '.csv', 'json': '.json', 'excel': '.xlsx'}[self.format_var.get()]
            out_path = get_output_path(Path(filename).stem + ext)
            self.output_file.set(str(out_path))
    
    def browse_output(self):
        ext = {'csv': '.csv', 'json': '.json', 'excel': '.xlsx'}[self.format_var.get()]
        filename = filedialog.asksaveasfilename(
            title="Save as",
            initialdir=str(ensure_output_dir()),
            defaultextension=ext,
            filetypes=[
                ("CSV files", "*.csv"),
                ("JSON files", "*.json"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.output_file.set(filename)
    
    def convert(self):
        vcd = self.vcd_file.get()
        output = self.output_file.get()
        
        if not vcd:
            messagebox.showerror("Error", "Please select a VCD file")
            return
        
        if not output:
            ext = {'csv': '.csv', 'json': '.json', 'excel': '.xlsx'}[self.format_var.get()]
            output = str(get_output_path(Path(vcd).stem + ext))
            self.output_file.set(output)
        
        self.status.config(text="Converting...", foreground="blue")
        self.root.update()
        
        try:
            signals, changes, metadata = parse_vcd(vcd)
            
            if not signals:
                messagebox.showerror("Error", "No signals found in VCD file")
                return
            
            rows = build_timeline(signals, changes)
            fmt = self.format_var.get()
            
            if fmt == 'csv':
                count = export_csv(signals, rows, output, self.time_unit.get(), self.value_fmt.get())
            elif fmt == 'json':
                count = export_json(signals, rows, output, metadata, self.time_unit.get(), self.value_fmt.get())
            elif fmt == 'excel':
                count = export_excel(signals, rows, output, self.time_unit.get(), self.value_fmt.get())
            
            self.status.config(text=f"Done! {len(signals)} signals, {count} rows", foreground="green")
            messagebox.showinfo("Success", f"Converted {len(signals)} signals, {count} rows\n\nSaved to:\n{output}")
            
        except Exception as e:
            self.status.config(text="Error", foreground="red")
            messagebox.showerror("Error", str(e))
    
    def run(self):
        self.root.mainloop()


# ============================================================
# Command Line
# ============================================================

def cli_convert(args):
    """Command line conversion."""
    
    if not args:
        print(__doc__)
        return
    
    vcd_file = args[0]
    output_file = None
    fmt = 'csv'
    value_fmt = 'hex'
    time_unit = 'us'
    
    i = 1
    while i < len(args):
        arg = args[i]
        if arg == '-o' and i + 1 < len(args):
            output_file = args[i + 1]
            i += 2
        elif arg == '--csv':
            fmt = 'csv'
            i += 1
        elif arg == '--json':
            fmt = 'json'
            i += 1
        elif arg == '--excel':
            fmt = 'excel'
            i += 1
        elif arg == '--hex':
            value_fmt = 'hex'
            i += 1
        elif arg == '--int' or arg == '--unsigned':
            value_fmt = 'int'
            i += 1
        elif arg == '--signed':
            value_fmt = 'signed'
            i += 1
        elif arg == '--smag':
            value_fmt = 'smag'
            i += 1
        elif arg == '--bin':
            value_fmt = 'bin'
            i += 1
        elif arg in ('--ps', '--ns', '--us', '--ms'):
            time_unit = arg[2:]
            i += 1
        else:
            if arg.endswith(('.csv', '.json', '.xlsx')):
                output_file = arg
            i += 1
    
    # Default output filename in converted_output folder
    if not output_file:
        ext = {'csv': '.csv', 'json': '.json', 'excel': '.xlsx'}[fmt]
        output_file = str(get_output_path(Path(vcd_file).stem + ext))
    
    print(f"Reading: {vcd_file}")
    signals, changes, metadata = parse_vcd(vcd_file)
    
    if not signals:
        print("Error: No signals found")
        return
    
    print(f"Found {len(signals)} signals, {len(changes)} value changes")
    rows = build_timeline(signals, changes)
    
    print(f"Writing: {output_file}")
    
    if fmt == 'csv':
        count = export_csv(signals, rows, output_file, time_unit, value_fmt)
    elif fmt == 'json':
        count = export_json(signals, rows, output_file, metadata, time_unit, value_fmt)
    elif fmt == 'excel':
        count = export_excel(signals, rows, output_file, time_unit, value_fmt)
    
    print(f"Done. {count} rows written.")


# ============================================================
# Main
# ============================================================

def main():
    args = sys.argv[1:]
    
    # Launch GUI if no args or --gui flag
    if not args or '--gui' in args:
        if HAS_TK:
            app = ConverterGUI()
            app.run()
        else:
            print("GUI requires tkinter. Use command line instead.")
            print(__doc__)
    else:
        cli_convert(args)


if __name__ == "__main__":
    main()
